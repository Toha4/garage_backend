import math
import os
from datetime import datetime

from django.conf import settings

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.helpers.reports import ALIGNMENT_CENTER
from app.helpers.reports import ALIGNMENT_LEFT
from app.helpers.reports import get_module_path
from app.helpers.reports import set_border
from app.helpers.reports import set_cell

from ..constants import COMPLETED
from ..models import Order


class OrderExcelCreator:
    def __init__(self, order: Order) -> None:
        self.__template_name = "order_template.xlsx"
        self.__order = order

    def __call__(self):
        wb = load_workbook(f"{get_module_path(__file__)}/templates/{self.__template_name}")
        ws = wb.active

        self.__print_main_data(ws)

        row_number = 12
        row_number = self.__print_works(ws, row_number)
        row_number += 1
        row_number = self.__print_materials(ws, row_number)
        row_number += 1
        self.__print_signatures(ws, row_number)

        path = f"{settings.MEDIA_ROOT}/temp"
        today = datetime.today()

        order_number = "__" if self.__order else self.__order.number
        filename = f"Заказ-наряд №{order_number} {today.strftime('%d%m%Y%H%M%S')}.xlsx"

        if not os.path.exists(path):
            os.makedirs(path)

        file_path = f"{path}/{filename}"
        wb.save(file_path)

        return f"{settings.MEDIA_URL}temp/{filename}"

    def __print_main_data(self, ws: Worksheet):
        order = self.__order

        ws["F1"] = "____" if order else order.number

        if order and order.date_begin:
            ws["A4"] = order.date_begin.strftime("%d.%m.%Y %H:%M")

        if order and order.date_end:
            ws["D4"] = order.date_end.strftime("%d.%m.%Y %H:%M")

        if order and order.reason:
            ws["G4"] = order.reason.name

        if order and order.car:
            ws["A7"] = order.car.state_number

        if order and order.car:
            ws["D7"] = order.car.name

        if order and order.odometer:
            ws["J7"] = order.odometer

        if order and order.post:
            ws["A10"] = order.post.name

        if order and order.note:
            ws["D10"] = order.note

    def __print_works(self, ws: Worksheet, row_start: int) -> int:
        row_number = row_start
        order_works = []
        is_complited = self.__order and self.__order.status == COMPLETED

        if self.__order:
            order_works = self.__order.order_works.all()

            if is_complited and not order_works:
                return row_number

        # Title table
        ws.row_dimensions[row_start].height = 25

        ws.merge_cells(f"A{row_number}:D{row_number}")
        set_cell(ws[f"A{row_number}"], "Работы", alignment=ALIGNMENT_CENTER, font=Font(bold=True))

        set_cell(ws[f"E{row_number}"], "Кол-во", alignment=ALIGNMENT_CENTER, font=Font(bold=True))

        ws.merge_cells(f"F{row_number}:G{row_number}")
        set_cell(ws[f"F{row_number}"], "Затраченно времени", alignment=ALIGNMENT_CENTER, font=Font(bold=True))

        ws.merge_cells(f"H{row_number}:K{row_number}")
        set_cell(ws[f"H{row_number}"], "Исполнитель", alignment=ALIGNMENT_CENTER, font=Font(bold=True))
        row_number += 1

        for work in order_works:
            self.__print_work_row(
                ws,
                row_number,
                work.work.name,
                str(work.quantity),
                self.__time_minutes_formated(work.time_minutes),
                self.__mechanics_formated(work.mechanics.all()),
            )
            row_number += 1

        if not is_complited and len(order_works) < 10:
            for _ in range(10 - len(order_works)):
                self.__print_work_row(ws, row_number, is_empty=True)
                row_number += 1

        set_border(ws, f"A{row_start}:K{row_number - 1}")

        return row_number

    def __print_work_row(self, ws: Worksheet, row: int, name="", quantity="", time="", mechanics="", is_empty=False):
        if is_empty:
            ws.row_dimensions[row].height = 18

        ws.merge_cells(f"A{row}:D{row}")
        set_cell(ws[f"A{row}"], name, alignment=ALIGNMENT_LEFT, wrap_text=True)

        set_cell(ws[f"E{row}"], quantity, alignment=ALIGNMENT_LEFT)

        ws.merge_cells(f"F{row}:G{row}")
        set_cell(ws[f"F{row}"], time, alignment=ALIGNMENT_LEFT)

        ws.merge_cells(f"H{row}:K{row}")
        set_cell(ws[f"H{row}"], mechanics, alignment=ALIGNMENT_LEFT, wrap_text=True)

        # For auto width (auto width not working with merge cells)
        set_cell(ws[f"N{row}"], name, alignment=ALIGNMENT_LEFT, wrap_text=True)
        set_cell(ws[f"O{row}"], mechanics, alignment=ALIGNMENT_LEFT, wrap_text=True)

    def __time_minutes_formated(self, tyme_minutes: int) -> str:
        time_str = ""

        if not tyme_minutes:
            return time_str

        minutes = math.floor(tyme_minutes % 60)
        hours = math.floor((tyme_minutes / 60) % 8)
        days = math.floor(tyme_minutes / 480)

        if days:
            time_str += f"{days}д "

        if hours:
            time_str += f"{hours}ч "

        if minutes:
            time_str += f"{minutes}м"

        return time_str

    def __mechanics_formated(self, mechanics) -> str:
        mechanics_str = ""

        if not mechanics:
            return mechanics_str

        for mechanic in mechanics:
            if mechanics_str != "":
                mechanics_str += "\n"

            mechanics_str += f"{mechanic.mechanic.short_fio} {'-' if mechanic.time_minutes else '' } {self.__time_minutes_formated(mechanic.time_minutes)}"

        return mechanics_str

    def __print_materials(self, ws: Worksheet, row_start: int) -> int:
        # TODO: Update method after added Materials application
        row_number = row_start

        if self.__order and self.__order.status == COMPLETED:
            return row_number

        # Title table
        ws.merge_cells(f"A{row_number}:G{row_number}")
        set_cell(ws[f"A{row_number}"], "Израсходованные материалы", alignment=ALIGNMENT_CENTER, font=Font(bold=True))

        ws.merge_cells(f"H{row_number}:I{row_number}")
        set_cell(ws[f"H{row_number}"], "Кол-во", alignment=ALIGNMENT_CENTER, font=Font(bold=True))

        ws.merge_cells(f"J{row_number}:K{row_number}")
        set_cell(ws[f"J{row_number}"], "Склад", alignment=ALIGNMENT_CENTER, font=Font(bold=True))
        row_number += 1

        for _ in range(10):
            self.__print_material_row(ws, row_number, is_empty=True)
            row_number += 1

        set_border(ws, f"A{row_start}:K{row_number - 1}")

        return row_number

    def __print_material_row(self, ws: Worksheet, row="", name="", quantity="", store_house="", is_empty=False):
        if is_empty:
            ws.row_dimensions[row].height = 18

        ws.merge_cells(f"A{row}:G{row}")
        set_cell(ws[f"A{row}"], name, alignment=ALIGNMENT_LEFT, wrap_text=True)

        ws.merge_cells(f"H{row}:I{row}")
        set_cell(ws[f"H{row}"], quantity, alignment=ALIGNMENT_LEFT)

        ws.merge_cells(f"J{row}:K{row}")
        set_cell(ws[f"J{row}"], store_house, alignment=ALIGNMENT_LEFT)

        # For auto width (auto width not working with merge cells)
        set_cell(ws[f"P{row}"], name, alignment=ALIGNMENT_LEFT, wrap_text=True)

    def __print_signatures(self, ws: Worksheet, row_start: int) -> int:
        empty_sign = "_____________________   _________"

        row_number = row_start

        ws[f"A{row_number}"] = "Ответсвенный:"
        row_number += 1
        if self.__order and self.__order.responsible:
            set_cell(ws[f"A{row_number}"], f"{self.__order.responsible.short_fio}  _________", font=Font(bold=True))
        else:
            ws[f"A{row_number}"] = empty_sign
        row_number += 2

        ws[f"A{row_number}"] = "Водитель:"
        row_number += 1
        if self.__order and self.__order.driver:
            set_cell(ws[f"A{row_number}"], f"{self.__order.driver.short_fio}  _________", font=Font(bold=True))
        else:
            ws[f"A{row_number}"] = empty_sign

        row_number = row_start
        ws[f"H{row_number}"] = "Исполнители:"
        row_number += 1

        mechanics = self.__get_works_mechanics()
        for mechanic in mechanics:
            set_cell(ws[f"H{row_number}"], f"{mechanic}  _________", font=Font(bold=True))
            row_number += 1

        is_complited = self.__order and self.__order.status == COMPLETED
        if not is_complited and len(mechanics) < 4:
            for _ in range(4 - len(mechanics)):
                ws[f"H{row_number}"] = empty_sign
                row_number += 1

        return row_number

    def __get_works_mechanics(self) -> list:
        mechanics = []

        if not self.__order:
            return mechanics

        for order_works in self.__order.order_works.all():
            for mechanic in order_works.mechanics.all():
                short_fio = mechanic.mechanic.short_fio
                if short_fio not in mechanics:
                    mechanics.append(short_fio)

        return mechanics
