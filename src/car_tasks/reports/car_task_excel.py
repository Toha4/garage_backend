import os
from copy import copy
from datetime import datetime

from django.conf import settings

from openpyxl import load_workbook

from app.helpers.reports import ALIGNMENT_LEFT
from app.helpers.reports import BOLD_FONT
from app.helpers.reports import get_module_path
from app.helpers.reports import set_border
from app.helpers.reports import set_cell


def car_tasks_excel(data):
    wb = load_workbook(f"{get_module_path(__file__)}/" f"templates/car_tasks_template.xlsx")
    ws = wb.active

    today = datetime.today()

    ws.merge_cells("A1:E1")
    set_cell(
        ws["A1"],
        f"Список задач на {today.strftime('%d.%m.%Y')}",
        font=BOLD_FONT,
        alignment=ALIGNMENT_LEFT,
    )

    start_row_number = 3
    row_number = copy(start_row_number)

    for index, row in enumerate(data):
        set_cell(ws[f"A{row_number}"], f"{index + 1}", number_format="0", alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"B{row_number}"], row["created"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"C{row_number}"], row["car_name"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"D{row_number}"], row["description"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"E{row_number}"], row["materials"], alignment=ALIGNMENT_LEFT)

        row_number += 1

    set_border(ws, f"A{start_row_number}:E{row_number - 1}")

    path = f"{settings.MEDIA_ROOT}/temp"

    filename = f"Список задач {today.strftime('%d%m%Y%H%M%S')}.xlsx"

    if not os.path.exists(path):
        os.makedirs(path)

    file_path = f"{path}/{filename}"
    wb.save(file_path)

    return f"{settings.MEDIA_URL}temp/{filename}"
