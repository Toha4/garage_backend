import os
from copy import copy
from datetime import datetime

from django.conf import settings

from openpyxl import load_workbook

from app.helpers.reports import ALIGNMENT_LEFT
from app.helpers.reports import BOLD_FONT
from app.helpers.reports import get_module_path
from app.helpers.reports import set_border
from app.helpers.reports import set_cell
from orders.helpers.time_minutes_formated import time_minutes_formated


def report_mechanics_excel(data, date_begin_begin, date_begin_end):
    wb = load_workbook(f"{get_module_path(__file__)}/" f"templates/report_mechanics_template.xlsx")
    ws = wb.active

    period = ""
    if date_begin_begin and date_begin_end:
        period = f"Отчет по работникам (слесарям) за период с {date_begin_begin} по {date_begin_end}"

    ws.merge_cells("A1:F1")
    set_cell(
        ws["A1"],
        period,
        font=BOLD_FONT,
        alignment=ALIGNMENT_LEFT,
    )

    start_row_number = 3
    row_number = copy(start_row_number)

    for index, row in enumerate(data):
        set_cell(ws[f"A{row_number}"], f"{index + 1}", number_format="0", alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"B{row_number}"], row["short_fio"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"C{row_number}"], time_minutes_formated(row["work_minutes_total"]), alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"D{row_number}"], row["repaired_cars_total"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"E{row_number}"], "; ".join(row["note_list"]), alignment=ALIGNMENT_LEFT)

        row_number += 1

    set_border(ws, f"A{start_row_number}:E{row_number - 1}")

    path = f"{settings.MEDIA_ROOT}/temp"
    today = datetime.today()

    filename = f"Отчет по работникам {today.strftime('%d%m%Y%H%M%S')}.xlsx"

    if not os.path.exists(path):
        os.makedirs(path)

    file_path = f"{path}/{filename}"
    wb.save(file_path)

    return f"{settings.MEDIA_URL}temp/{filename}"
