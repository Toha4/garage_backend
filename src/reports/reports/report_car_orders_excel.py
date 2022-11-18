import os
from copy import copy
from datetime import datetime

from django.conf import settings

from openpyxl import load_workbook

from app.helpers.reports import ALIGNMENT_LEFT
from app.helpers.reports import BOLD_FONT
from app.helpers.reports import get_module_path
from app.helpers.reports import set_border
from app.helpers.reports import set_cell
from core.models import Car
from orders.helpers.time_minutes_formated import time_minutes_formated


def report_car_orders_excel(car_pk, data, date_begin_begin, date_begin_end):
    wb = load_workbook(f"{get_module_path(__file__)}/" f"templates/report_car_orders_template.xlsx")
    ws = wb.active

    car = Car.objects.get(pk=car_pk)
    car_name = f"{car.name} гос. № {car.state_number}"

    period = f"Заказ-наряды по {car_name}"
    if date_begin_begin and date_begin_end:
        period += f" за период с {date_begin_begin} по {date_begin_end}"

    ws.merge_cells("A1:H1")
    set_cell(
        ws["A1"],
        period,
        font=BOLD_FONT,
        alignment=ALIGNMENT_LEFT,
    )

    start_row_number = 3
    row_number = copy(start_row_number)

    total_time_mnutes = 0
    total_sum = 0.0

    for index, row in enumerate(data):
        set_cell(ws[f"A{row_number}"], f"{index + 1}", number_format="0", alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"B{row_number}"], row["date_begin"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"C{row_number}"], f'№{row["number"]}', alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"D{row_number}"], row["reason_name"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"E{row_number}"], ";\n".join(f"- {work}" for work in row["work_list"]), alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"F{row_number}"], time_minutes_formated(row["work_minutes_total"]), alignment=ALIGNMENT_LEFT)
        set_cell(
            ws[f"G{row_number}"],
            ";\n".join(f"- {material}" for material in row["material_list"]),
            alignment=ALIGNMENT_LEFT,
        )
        set_cell(ws[f"H{row_number}"], row["sum_total"], number_format="#,##0.00", alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"I{row_number}"], row["note"], alignment=ALIGNMENT_LEFT)

        total_time_mnutes += row["work_minutes_total"]
        total_sum += float(row["sum_total"])

        row_number += 1

    total_row_number = copy(row_number)

    set_cell(ws[f"B{total_row_number}"], "Итого:", font=BOLD_FONT, alignment=ALIGNMENT_LEFT)

    set_cell(
        ws[f"F{total_row_number}"],
        time_minutes_formated(total_time_mnutes),
        font=BOLD_FONT,
        alignment=ALIGNMENT_LEFT,
    )

    set_cell(
        ws[f"H{total_row_number}"],
        total_sum,
        font=BOLD_FONT,
        number_format="#,##0.00",
        alignment=ALIGNMENT_LEFT,
    )

    set_border(ws, f"A{start_row_number}:I{total_row_number}")

    path = f"{settings.MEDIA_ROOT}/temp"
    today = datetime.today()

    filename = f"Отчет по заказ-нарядам {car_name} {today.strftime('%d%m%Y%H%M%S')}.xlsx"

    if not os.path.exists(path):
        os.makedirs(path)

    file_path = f"{path}/{filename}"
    wb.save(file_path)

    return f"{settings.MEDIA_URL}temp/{filename}"
