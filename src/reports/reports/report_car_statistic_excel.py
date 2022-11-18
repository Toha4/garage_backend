import os
from copy import copy
from datetime import datetime

from django.conf import settings

from openpyxl import load_workbook

from app.helpers.reports import ALIGNMENT_CENTER
from app.helpers.reports import ALIGNMENT_LEFT
from app.helpers.reports import BOLD_FONT
from app.helpers.reports import get_module_path
from app.helpers.reports import set_border
from app.helpers.reports import set_cell
from core.models import Car


def report_car_statistic_excel(car_pk, data, date_begin_begin, date_begin_end):
    wb = load_workbook(f"{get_module_path(__file__)}/" f"templates/report_car_statistic_template.xlsx")
    ws = wb.active

    car = Car.objects.get(pk=car_pk)

    period = f"Статистика по {car.name} гос. № {car.state_number}"
    if date_begin_begin and date_begin_end:
        period += f" за период с {date_begin_begin} по {date_begin_end}"

    ws.merge_cells("A1:E1")
    set_cell(
        ws["A1"],
        period,
        font=BOLD_FONT,
        alignment=ALIGNMENT_LEFT,
    )

    start_row_number = 3
    row_number = copy(start_row_number)

    for row in data:
        set_cell(ws[f"A{row_number}"], row["name_param"], alignment=ALIGNMENT_LEFT)
        set_cell(ws[f"B{row_number}"], row["total"], alignment=ALIGNMENT_CENTER)
        set_cell(ws[f"C{row_number}"], row["maintenance"], alignment=ALIGNMENT_CENTER)
        set_cell(ws[f"D{row_number}"], row["repair"], alignment=ALIGNMENT_CENTER)
        set_cell(ws[f"E{row_number}"], row["other"], alignment=ALIGNMENT_CENTER)

        row_number += 1

    set_border(ws, f"A{start_row_number}:E{row_number - 1}")

    path = f"{settings.MEDIA_ROOT}/temp"
    today = datetime.today()

    filename = f"Статистика по ТС {today.strftime('%d%m%Y%H%M%S')}.xlsx"

    if not os.path.exists(path):
        os.makedirs(path)

    file_path = f"{path}/{filename}"
    wb.save(file_path)

    return f"{settings.MEDIA_URL}temp/{filename}"
